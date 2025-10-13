#!/usr/bin/env python3
"""
Clean the ad-hoc "daily eggs report.xlsx" spreadsheet into import-ready CSV and SQL files.

Outputs are written to ``db_snapshots/import_ready`` and include:
  * daily_reports_clean.csv         – normalized daily production data
  * daily_report_notes.csv          – free-form rows / comments from the sheet
  * egg_sales_clean.csv             – normalized outgoing sales data
  * daily_reports_insert.sql        – INSERT statements for daily production
  * egg_sales_insert.sql            – INSERT statements for sales
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_PATH = BASE_DIR / "daily eggs report.xlsx"
OUTPUT_DIR = BASE_DIR / "db_snapshots" / "import_ready"
EGGS_PER_TRAY = Decimal("30")
QUANTIZE_6 = Decimal("0.000001")


def parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def to_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def decimal_to_str(value: Optional[Decimal]) -> Optional[str]:
    if value is None:
        return None
    formatted = format(value, "f").rstrip("0").rstrip(".")
    return formatted or "0"


def quantize_decimal(value: Decimal, places: int = 6) -> Decimal:
    if places <= 0:
        return value
    quant = Decimal(1) / (Decimal(10) ** places)
    return value.quantize(quant, rounding=ROUND_HALF_UP)


def safe_decimal(value: Optional[Any]) -> Decimal:
    if value in (None, "", "NULL"):
        return Decimal(0)
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return Decimal(0)
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal(0)


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, str):
        return "'" + value.replace("'", "''") + "'"
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    if isinstance(value, Decimal):
        return decimal_to_str(value) or "0"
    if isinstance(value, date):
        return "'" + value.isoformat() + "'"
    return "'" + str(value).replace("'", "''") + "'"


def normalise_daily_reports(wb) -> Dict[str, Sequence[Dict[str, Any]]]:
    sheet = wb["Daily reports "]
    header_row = next(sheet.iter_rows(values_only=True))
    headers = [h.strip() if isinstance(h, str) else h for h in header_row]
    col_map = {
        "Date": "date",
        "Age in weeks": "age_weeks",
        "No. of Eggs Picked": "eggs_picked",
        "No. of trays picked": "trays_picked",
        "Damages": "damages",
        "No. of Trays out": "trays_out",
        "No. of Trays in": "trays_in",
        "Extras": "extras",
        "Motarlity": "mortality",
        "No. of bags of feed used": "feed_bags",
        "No. of Birds": "bird_count",
        "Daily P/percenage": "production_pct",
    }

    daily_records: List[Dict[str, Any]] = []
    note_records: List[Dict[str, Any]] = []
    current_age: Optional[int] = None

    for excel_row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        mapped: Dict[str, Any] = {}
        for header, value in zip(headers, row):
            alias = col_map.get(header)
            if alias:
                mapped[alias] = value

        if not any(v not in (None, "") for v in mapped.values()):
            continue

        raw_date = mapped.get("date")
        parsed_date = parse_date(raw_date)
        extras_raw = mapped.get("extras")

        if parsed_date is None:
            payload = {
                key: value
                for key, value in mapped.items()
                if value not in (None, "")
            }
            if payload:
                note_records.append(
                    {
                        "sheet": "daily_reports",
                        "row_index": excel_row_idx,
                        "note": f"Non-dated row with data: {payload}",
                    }
                )
            continue

        age_value = mapped.get("age_weeks")
        parsed_age = to_int(age_value)
        if parsed_age is not None:
            current_age = parsed_age
        age_weeks = current_age

        eggs_picked = to_int(mapped.get("eggs_picked"))
        trays_picked = to_int(mapped.get("trays_picked"))
        damages = to_int(mapped.get("damages"))
        trays_out = to_int(mapped.get("trays_out"))
        trays_in = to_int(mapped.get("trays_in"))
        mortality = to_int(mapped.get("mortality"))

        feed_raw = mapped.get("feed_bags")
        feed_used = to_int(feed_raw)
        feed_used = abs(feed_used) if feed_used is not None else None

        bird_count = to_int(mapped.get("bird_count"))
        production_pct = to_decimal(mapped.get("production_pct"))

        extras_count: Optional[int] = None
        notes_text: Optional[str] = None
        if isinstance(extras_raw, (int, float)):
            extras_count = int(extras_raw)
        elif extras_raw not in (None, ""):
            notes_text = str(extras_raw).strip()

        daily_records.append(
            {
                "report_date": parsed_date.isoformat(),
                "age_weeks": age_weeks,
                "eggs_picked": eggs_picked,
                "trays_picked": trays_picked,
                "damages": damages,
                "trays_out": trays_out,
                "trays_in": trays_in,
                "extras_count": extras_count,
                "mortality": mortality,
                "feed_bags_used": feed_used,
                "bird_count": bird_count,
                "production_percent": (
                    float(production_pct) if production_pct is not None else None
                ),
                "notes": notes_text,
            }
        )

    daily_records.sort(key=lambda row: row["report_date"])

    if daily_records:
        note_records.append(
            {
                "sheet": "daily_reports",
                "row_index": None,
                "note": "Feed usage values were stored as negatives in the sheet; "
                "the script now records their absolute values.",
            }
        )

    return {"daily_records": daily_records, "note_records": note_records}


def normalise_sales(wb) -> Dict[str, Any]:
    sheet = wb["Eggs Sales"]
    header_row = next(sheet.iter_rows(values_only=True))
    headers = [h.strip() if isinstance(h, str) else h for h in header_row]
    col_map = {
        "Date": "date",
        "Customer Name": "customer",
        "No. of trays": "trays",
        "Price /tray": "price_per_tray",
        "Total Amount": "total_amount",
        "Invoice No.": "invoice",
        "Payment Received": "payment_received",
        "Balance": "balance",
    }

    sales_records: List[Dict[str, Any]] = []
    note_records: List[Dict[str, Any]] = []

    for excel_row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        mapped: Dict[str, Any] = {}
        for header, value in zip(headers, row):
            alias = col_map.get(header)
            if alias:
                mapped[alias] = value

        if not any(v not in (None, "") for v in mapped.values()):
            continue

        parsed_date = parse_date(mapped.get("date"))
        if parsed_date is None:
            note_records.append(
                {
                    "sheet": "egg_sales",
                    "row_index": excel_row_idx,
                    "note": f"Skipped row with unparsed date: {mapped}",
                }
            )
            continue

        customer = (mapped.get("customer") or "").strip() or None
        trays = to_int(mapped.get("trays"))
        price = to_decimal(mapped.get("price_per_tray"))
        total_amount = to_decimal(mapped.get("total_amount"))
        payment = to_decimal(mapped.get("payment_received"))
        balance = to_decimal(mapped.get("balance"))

        computed_total = (
            Decimal(trays) * price if trays is not None and price is not None else None
        )

        final_total = computed_total if computed_total is not None else total_amount

        computed_balance: Optional[Decimal] = None
        if final_total is not None and payment is not None:
            computed_balance = final_total - payment
        elif final_total is not None and payment is None:
            computed_balance = final_total
        final_balance = computed_balance if computed_balance is not None else balance

        note_text: Optional[str] = None
        if (
            total_amount is not None
            and computed_total is not None
            and total_amount != computed_total
        ):
            note_text = (
                f"Total mismatch (sheet={decimal_to_str(total_amount)}, "
                f"computed={decimal_to_str(computed_total)})"
            )

        sales_records.append(
            {
                "sale_date": parsed_date.isoformat(),
                "customer_name": customer,
                "trays": trays,
                "price_per_tray": (
                    float(price) if price is not None else None
                ),
                "total_amount": (
                    float(final_total) if final_total is not None else None
                ),
                "invoice_no": mapped.get("invoice"),
                "payment_received": (
                    float(payment) if payment is not None else None
                ),
                "balance": (
                    float(final_balance) if final_balance is not None else None
                ),
                "notes": note_text,
            }
        )

        if price is None and trays:
            note_records.append(
                {
                    "sheet": "egg_sales",
                    "row_index": excel_row_idx,
                    "note": "Price per tray missing; total/balance were inferred where possible.",
                }
            )

    sales_records.sort(key=lambda row: row["sale_date"])

    return {"sales_records": sales_records, "note_records": note_records}


def write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_sql(path: Path, table: str, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]):
    statements: List[str] = []
    for row in rows:
        values = [sql_literal(row.get(field)) for field in fieldnames]
        statements.append(
            f"INSERT INTO {table} ({', '.join(fieldnames)}) VALUES ({', '.join(values)});"
        )
    with path.open("w", encoding="utf-8") as handle:
        handle.write("\n".join(statements))


def build_production_description(record: Dict[str, Any]) -> str:
    parts: List[str] = []
    eggs_picked = record.get("eggs_picked")
    if eggs_picked:
        parts.append(f"Eggs collected: {eggs_picked}")
    extras = record.get("extras_count")
    if extras:
        parts.append(f"Loose eggs: {extras}")
    damages = record.get("damages")
    if damages:
        parts.append(f"Damaged eggs: {damages}")
    mortality = record.get("mortality")
    if mortality:
        parts.append(f"Bird mortality: {mortality}")
    feed = record.get("feed_bags_used")
    if feed:
        parts.append(f"Feed used: {feed} bags")
    age = record.get("age_weeks")
    if age:
        parts.append(f"Flock age: {age} weeks")
    production_pct = record.get("production_percent")
    if production_pct is not None:
        parts.append(f"Production %: {production_pct}")
    notes = record.get("notes")
    if notes:
        parts.append(f"Notes: {notes}")
    return "; ".join(str(p) for p in parts)


def generate_production_import_sql(path: Path, daily_records: Sequence[Dict[str, Any]]) -> None:
    total_feed_bags = Decimal(0)
    total_mortality = Decimal(0)
    for record in daily_records:
        total_feed_bags += safe_decimal(record.get("feed_bags_used"))
        total_mortality += safe_decimal(record.get("mortality"))

    production_rows: List[Dict[str, Any]] = []
    for record in daily_records:
        trays_picked = safe_decimal(record.get("trays_picked"))
        if trays_picked <= 0:
            continue

        damages_eggs = safe_decimal(record.get("damages"))
        extras_eggs = safe_decimal(record.get("extras_count"))

        damages_trays = Decimal(0)
        extras_trays = Decimal(0)
        if damages_eggs > 0:
            damages_trays = quantize_decimal(damages_eggs / EGGS_PER_TRAY)
        if extras_eggs > 0:
            extras_trays = quantize_decimal(extras_eggs / EGGS_PER_TRAY)

        description = build_production_description(record)

        production_rows.append(
            {
                "report_date": record["report_date"],
                "trays_picked": quantize_decimal(trays_picked),
                "damages_trays": quantize_decimal(damages_trays),
                "extras_trays": quantize_decimal(extras_trays),
                "description": description,
            }
        )

    if not production_rows:
        path.write_text("-- No production rows to import.\n", encoding="utf-8")
        return

    start_date = min(row["report_date"] for row in production_rows)
    end_date = max(row["report_date"] for row in production_rows)

    lines: List[str] = []
    lines.append("-- Auto-generated egg production import script")
    lines.append("-- Generated by scripts/clean_daily_eggs_report.py")
    lines.append("START TRANSACTION;")
    lines.append("")
    lines.append("SET @product_id := '60397309';")
    lines.append("SET @shed_id := (SELECT id FROM sheds ORDER BY id LIMIT 1);")
    lines.append("SET @stock_location_id := (SELECT id FROM stock_locations WHERE location_type = 'shed' ORDER BY id LIMIT 1);")
    lines.append("SET @unit_id := (SELECT unit_id FROM product_units WHERE product_id = @product_id AND is_base = 1 LIMIT 1);")
    lines.append("SET @creator_id := (SELECT id FROM users ORDER BY id LIMIT 1);")
    lines.append("SET @reason_output_id := (SELECT id FROM stock_movement_reasons WHERE reason_code = 'PRODUCTION_OUTPUT' LIMIT 1);")
    lines.append("SET @reason_loss_id := (SELECT id FROM stock_movement_reasons WHERE reason_code = 'PRODUCTION_LOSS' LIMIT 1);")
    lines.append("")
    lines.append("SET @shed_id := IFNULL(@shed_id, 1);")
    lines.append("SET @stock_location_id := IFNULL(@stock_location_id, 1);")
    lines.append("SET @unit_id := IFNULL(@unit_id, 1);")
    lines.append("SET @creator_id := IFNULL(@creator_id, 1);")
    lines.append("")
    lines.append(f"SET @import_start := DATE('{start_date}');")
    lines.append(f"SET @import_end := DATE('{end_date}');")
    lines.append("")
    lines.append("DROP TEMPORARY TABLE IF EXISTS tmp_existing_productions;")
    lines.append("CREATE TEMPORARY TABLE tmp_existing_productions (production_id INT PRIMARY KEY);")
    lines.append("INSERT INTO tmp_existing_productions (production_id)")
    lines.append("SELECT id FROM productions")
    lines.append("WHERE output_product_id = @product_id")
    lines.append("  AND DATE(created_at) BETWEEN @import_start AND @import_end;")
    lines.append("")
    lines.append("DELETE FROM stock_movements")
    lines.append("WHERE reference_type = 'production'")
    lines.append("  AND reference_id IN (SELECT production_id FROM tmp_existing_productions);")
    lines.append("")
    lines.append("DELETE FROM stock_lots")
    lines.append("WHERE source_type = 'production'")
    lines.append("  AND source_reference IN (SELECT production_id FROM tmp_existing_productions);")
    lines.append("")
    lines.append("DELETE FROM productions")
    lines.append("WHERE id IN (SELECT production_id FROM tmp_existing_productions);")
    lines.append("")
    lines.append("DROP TEMPORARY TABLE IF EXISTS tmp_existing_productions;")
    lines.append("")

    for index, row in enumerate(production_rows, start=1):
        report_date = row["report_date"]
        trays_picked = row["trays_picked"]
        damages_trays = row["damages_trays"]
        extras_trays = row["extras_trays"]
        description = row["description"]

        damages_str = decimal_to_str(damages_trays) or "0"
        extras_str = decimal_to_str(extras_trays) or "0"
        trays_str = decimal_to_str(trays_picked) or "0"
        quantity_in_str = decimal_to_str(trays_picked + extras_trays) or "0"
        loss_total = damages_trays
        loss_str = decimal_to_str(loss_total) or "0"

        lot_code = f"PRD-{report_date.replace('-', '')}-{index:02d}"
        production_name = f"Egg Production {report_date}"
        lot_notes = description
        narration_output = f"Production output {report_date}"
        narration_loss = f"Production loss {report_date}"

        lines.append(f"-- Production record for {report_date}")
        lines.append(
            "INSERT INTO productions "
            "(name, shed_id, stock_location_id, output_product_id, output_unit_id, description, "
            "produced_total_qty, produced_mortality_qty, produced_damaged_qty, produced_extras_qty, "
            "unit_type_id, created_at, updated_at)"
        )
        lines.append(
            "VALUES "
            f"({sql_literal(production_name)}, @shed_id, @stock_location_id, @product_id, @unit_id, "
            f"{sql_literal(description)}, {trays_str}, 0, {damages_str}, {extras_str}, @unit_id, "
            f"'{report_date} 06:00:00', '{report_date} 06:00:00');"
        )
        lines.append("SET @production_id := LAST_INSERT_ID();")
        lines.append("")

        lines.append(
            "INSERT INTO stock_lots "
            "(lot_code, product_id, base_unit_id, location_id, source_type, source_reference, "
            "source_reference_line, production_date, expiry_date, grade, initial_quantity, status, notes, "
            "created_by, created_at, updated_by, updated_at)"
        )
        lines.append(
            "VALUES "
            f"({sql_literal(lot_code)}, @product_id, @unit_id, @stock_location_id, 'production', "
            "@production_id, NULL, "
            f"DATE('{report_date}'), NULL, NULL, {trays_str}, 'open', {sql_literal(lot_notes)}, "
            f"@creator_id, '{report_date} 06:05:00', NULL, '{report_date} 06:05:00');"
        )
        lines.append("SET @lot_id := LAST_INSERT_ID();")
        lines.append("")

        lines.append(
            "INSERT INTO production_output_items "
            "(production_id, product_id, unit_id, location_id, lot_id, quantity, loss_quantity, "
            "cost_per_unit, cost_total, created_by, created_at, updated_by, updated_at)"
        )
        lines.append(
            "VALUES "
            f"(@production_id, @product_id, @unit_id, @stock_location_id, @lot_id, {trays_str}, "
            f"{loss_str}, 0, 0, @creator_id, '{report_date} 06:10:00', NULL, '{report_date} 06:10:00');"
        )
        lines.append("SET @output_item_id := LAST_INSERT_ID();")
        lines.append("")

        lines.append(
            "INSERT INTO stock_movements "
            "(movement_date, product_id, lot_id, location_id, unit_id, quantity_in, quantity_out, "
            "cost_per_unit, cost_total, reason_id, reference_type, reference_id, reference_line_id, "
            "narration, created_by, created_at, updated_by, updated_at)"
        )
        lines.append(
            "VALUES "
            f"('{report_date} 06:15:00', @product_id, @lot_id, @stock_location_id, @unit_id, "
            f"{quantity_in_str}, 0, 0, 0, @reason_output_id, 'production', @production_id, "
            f"@output_item_id, {sql_literal(narration_output)}, @creator_id, "
            f"'{report_date} 06:15:00', NULL, '{report_date} 06:15:00');"
        )
        lines.append("")

        if loss_total > 0:
            lines.append(
                "INSERT INTO stock_movements "
                "(movement_date, product_id, lot_id, location_id, unit_id, quantity_in, quantity_out, "
                "cost_per_unit, cost_total, reason_id, reference_type, reference_id, reference_line_id, "
                "narration, created_by, created_at, updated_by, updated_at)"
            )
            lines.append(
                "VALUES "
                f"('{report_date} 06:20:00', @product_id, @lot_id, @stock_location_id, @unit_id, "
                f"0, {loss_str}, 0, 0, @reason_loss_id, 'production', @production_id, @output_item_id, "
                f"{sql_literal(narration_loss)}, @creator_id, '{report_date} 06:20:00', NULL, "
                f"'{report_date} 06:20:00');"
            )
            lines.append("")

    total_feed_str = decimal_to_str(quantize_decimal(total_feed_bags, 2)) or "0"
    total_mortality_int = int(total_mortality)

    lines.append("-- Update livestock headcount based on Excel mortality totals")
    lines.append(f"SET @total_mortality := {total_mortality_int};")
    lines.append("SET @livestock_id := (SELECT id FROM livestocks ORDER BY id LIMIT 1);")
    lines.append("SET @livestock_base := (SELECT COALESCE(livestock_total_qty, 0) + COALESCE(livestock_mortality_qty, 0) FROM livestocks WHERE id = @livestock_id);")
    lines.append("SET @livestock_target := GREATEST(@livestock_base - @total_mortality, 0);")
    lines.append("UPDATE livestocks")
    lines.append("SET livestock_mortality_qty = @total_mortality,")
    lines.append("    livestock_total_qty = @livestock_target,")
    lines.append("    updated_at = NOW()")
    lines.append("WHERE @livestock_id IS NOT NULL AND id = @livestock_id;")
    lines.append("")

    lines.append("-- Update aggregated feed usage totals from Excel daily feed records")
    lines.append(f"SET @total_feed_used := {total_feed_str};")
    lines.append("SET @feed_usage_id := (SELECT id FROM feed_usages ORDER BY id LIMIT 1);")
    lines.append("UPDATE feed_usages")
    lines.append("SET total_purchased_qty = @total_feed_used,")
    lines.append("    used_total_qty = @total_feed_used,")
    lines.append("    total_wasted_qty = 0,")
    lines.append("    total_instock_qty = 0,")
    lines.append("    updated_at = NOW()")
    lines.append("WHERE @feed_usage_id IS NOT NULL AND id = @feed_usage_id;")
    lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Cannot locate spreadsheet at {EXCEL_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(EXCEL_PATH, data_only=True)

    daily_payload = normalise_daily_reports(workbook)
    sales_payload = normalise_sales(workbook)

    daily_records = list(daily_payload["daily_records"])
    daily_notes = list(daily_payload["note_records"])
    sales_records = list(sales_payload["sales_records"])
    sales_notes = list(sales_payload["note_records"])

    write_csv(
        OUTPUT_DIR / "daily_reports_clean.csv",
        [
            "report_date",
            "age_weeks",
            "eggs_picked",
            "trays_picked",
            "damages",
            "trays_out",
            "trays_in",
            "extras_count",
            "mortality",
            "feed_bags_used",
            "bird_count",
            "production_percent",
            "notes",
        ],
        daily_records,
    )

    write_csv(
        OUTPUT_DIR / "daily_report_notes.csv",
        ["sheet", "row_index", "note"],
        daily_notes + sales_notes,
    )

    write_csv(
        OUTPUT_DIR / "egg_sales_clean.csv",
        [
            "sale_date",
            "customer_name",
            "trays",
            "price_per_tray",
            "total_amount",
            "invoice_no",
            "payment_received",
            "balance",
            "notes",
        ],
        sales_records,
    )

    write_sql(
        OUTPUT_DIR / "daily_reports_insert.sql",
        "staging_daily_reports",
        [
            "report_date",
            "age_weeks",
            "eggs_picked",
            "trays_picked",
            "damages",
            "trays_out",
            "trays_in",
            "extras_count",
            "mortality",
            "feed_bags_used",
            "bird_count",
            "production_percent",
            "notes",
        ],
        daily_records,
    )

    write_sql(
        OUTPUT_DIR / "egg_sales_insert.sql",
        "staging_egg_sales",
        [
            "sale_date",
            "customer_name",
            "trays",
            "price_per_tray",
            "total_amount",
            "invoice_no",
            "payment_received",
            "balance",
            "notes",
        ],
        sales_records,
    )

    generate_production_import_sql(
        OUTPUT_DIR / "daily_productions_import.sql",
        daily_records,
    )

    print("Daily production rows:", len(daily_records))
    print("Sales rows:", len(sales_records))
    print("Notes captured:", len(daily_notes) + len(sales_notes))
    print(f"Output directory: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
